import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import threading

class SimpleExcelMerger:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Birleştirme Uygulaması")
        self.root.geometry("900x700")
        self.root.configure(bg='#f0f0f0')
        
        # Dosya yolları
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.output_path = tk.StringVar()
        
        # Sütun seçimi
        self.file1_columns = []
        self.file2_columns = []
        self.file1_column_var = tk.StringVar()
        self.file2_column_var = tk.StringVar()
        
        # İlerleme durumu
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value="Hazır")
        
        self.setup_ui()
        
    def setup_ui(self):
        # Ana konteyner
        main_frame = tk.Frame(self.root, bg='#f0f0f0', padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Başlık
        title_label = tk.Label(main_frame, text="Excel Dosya Birleştirme Uygulaması", 
                              font=('Arial', 18, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        title_label.pack(pady=(0, 10))
        
        # Açıklama
        desc_label = tk.Label(main_frame, 
                             text="Seçilen sütunlara göre iki Excel dosyasını birleştirir ve eksik verileri doldurur",
                             font=('Arial', 11), bg='#f0f0f0', fg='#7f8c8d')
        desc_label.pack(pady=(0, 20))
        
        # Dosya seçim bölümü
        self.create_file_selection_section(main_frame)
        
        # Sütun seçim bölümü
        self.create_column_selection_section(main_frame)
        
        # Çıktı ayarları
        self.create_output_section(main_frame)
        
        # İşlem butonları
        self.create_action_buttons(main_frame)
        
        # İlerleme ve durum bölümü
        self.create_progress_section(main_frame)
        
        # Sonuç bölümü
        self.create_results_section(main_frame)
        
    def create_file_selection_section(self, parent):
        # Dosya seçim çerçevesi
        file_frame = tk.LabelFrame(parent, text="📁 Dosya Seçimi", font=('Arial', 12, 'bold'),
                                  bg='#f0f0f0', fg='#2c3e50', padx=15, pady=15)
        file_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Dosya 1
        file1_frame = tk.Frame(file_frame, bg='#f0f0f0')
        file1_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(file1_frame, text="1. Excel Dosyası:", font=('Arial', 11, 'bold'),
                bg='#f0f0f0', fg='#2c3e50').pack(anchor=tk.W)
        
        file1_input_frame = tk.Frame(file1_frame, bg='#f0f0f0')
        file1_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.file1_entry = tk.Entry(file1_input_frame, textvariable=self.file1_path, 
                                   font=('Arial', 10), state='readonly', width=60)
        self.file1_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        file1_btn = tk.Button(file1_input_frame, text="📂 Dosya Seç", 
                             command=lambda: self.select_file(1),
                             bg='#3498db', fg='white', font=('Arial', 9, 'bold'),
                             relief=tk.FLAT, padx=15, pady=5)
        file1_btn.pack(side=tk.RIGHT)
        
        # Dosya 2
        file2_frame = tk.Frame(file_frame, bg='#f0f0f0')
        file2_frame.pack(fill=tk.X)
        
        tk.Label(file2_frame, text="2. Excel Dosyası:", font=('Arial', 11, 'bold'),
                bg='#f0f0f0', fg='#2c3e50').pack(anchor=tk.W)
        
        file2_input_frame = tk.Frame(file2_frame, bg='#f0f0f0')
        file2_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.file2_entry = tk.Entry(file2_input_frame, textvariable=self.file2_path, 
                                   font=('Arial', 10), state='readonly', width=60)
        self.file2_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        file2_btn = tk.Button(file2_input_frame, text="📂 Dosya Seç", 
                             command=lambda: self.select_file(2),
                             bg='#3498db', fg='white', font=('Arial', 9, 'bold'),
                             relief=tk.FLAT, padx=15, pady=5)
        file2_btn.pack(side=tk.RIGHT)
        
    def create_column_selection_section(self, parent):
        # Sütun seçim çerçevesi
        column_frame = tk.LabelFrame(parent, text="📋 Birleştirme Sütunu Seçimi", font=('Arial', 12, 'bold'),
                                    bg='#f0f0f0', fg='#2c3e50', padx=15, pady=15)
        column_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Açıklama
        info_label = tk.Label(column_frame, 
                             text="Her iki dosyada da bulunan ve birleştirme için kullanılacak sütunları seçin",
                             font=('Arial', 10, 'italic'), bg='#f0f0f0', fg='#7f8c8d')
        info_label.pack(pady=(0, 15))
        
        # Sütun seçim çerçeveleri
        columns_container = tk.Frame(column_frame, bg='#f0f0f0')
        columns_container.pack(fill=tk.X)
        
        # Dosya 1 sütun seçimi
        file1_col_frame = tk.Frame(columns_container, bg='#f0f0f0')
        file1_col_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(file1_col_frame, text="1. Dosya Sütunu:", font=('Arial', 11, 'bold'),
                bg='#f0f0f0', fg='#2c3e50').pack(anchor=tk.W)
        
        self.file1_column_combo = ttk.Combobox(file1_col_frame, textvariable=self.file1_column_var,
                                              state="readonly", font=('Arial', 10), width=30)
        self.file1_column_combo.pack(fill=tk.X, pady=(5, 0))
        
        # Dosya 2 sütun seçimi
        file2_col_frame = tk.Frame(columns_container, bg='#f0f0f0')
        file2_col_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
        
        tk.Label(file2_col_frame, text="2. Dosya Sütunu:", font=('Arial', 11, 'bold'),
                bg='#f0f0f0', fg='#2c3e50').pack(anchor=tk.W)
        
        self.file2_column_combo = ttk.Combobox(file2_col_frame, textvariable=self.file2_column_var,
                                              state="readonly", font=('Arial', 10), width=30)
        self.file2_column_combo.pack(fill=tk.X, pady=(5, 0))
        
        # Otomatik eşleştirme butonu
        auto_match_btn = tk.Button(column_frame, text="🔄 Otomatik Eşleştir", 
                                  command=self.auto_match_columns,
                                  bg='#9b59b6', fg='white', font=('Arial', 10, 'bold'),
                                  relief=tk.FLAT, padx=20, pady=8)
        auto_match_btn.pack(pady=(15, 0))
        
    def create_output_section(self, parent):
        # Çıktı ayarları
        output_frame = tk.LabelFrame(parent, text="💾 Çıktı Ayarları", font=('Arial', 12, 'bold'),
                                    bg='#f0f0f0', fg='#2c3e50', padx=15, pady=15)
        output_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(output_frame, text="Çıktı Dosyası:", font=('Arial', 11, 'bold'),
                bg='#f0f0f0', fg='#2c3e50').pack(anchor=tk.W)
        
        output_input_frame = tk.Frame(output_frame, bg='#f0f0f0')
        output_input_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.output_entry = tk.Entry(output_input_frame, textvariable=self.output_path, 
                                    font=('Arial', 10), width=60)
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        output_btn = tk.Button(output_input_frame, text="📁 Klasör Seç", 
                              command=self.select_output_file,
                              bg='#27ae60', fg='white', font=('Arial', 9, 'bold'),
                              relief=tk.FLAT, padx=15, pady=5)
        output_btn.pack(side=tk.RIGHT)
        
        # Varsayılan çıktı dosyası
        self.output_path.set("birlestirilmis_dosya.xlsx")
        
    def create_action_buttons(self, parent):
        # Buton çerçevesi
        button_frame = tk.Frame(parent, bg='#f0f0f0')
        button_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Birleştir butonu
        self.merge_button = tk.Button(button_frame, text="🔄 Dosyaları Birleştir", 
                                     command=self.merge_files,
                                     bg='#e74c3c', fg='white', font=('Arial', 13, 'bold'),
                                     relief=tk.FLAT, padx=30, pady=12)
        self.merge_button.pack(side=tk.LEFT, padx=(0, 15))
        
        # Temizle butonu
        clear_btn = tk.Button(button_frame, text="🗑️ Temizle", 
                             command=self.clear_all,
                             bg='#95a5a6', fg='white', font=('Arial', 11, 'bold'),
                             relief=tk.FLAT, padx=25, pady=12)
        clear_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # Çıkış butonu
        exit_btn = tk.Button(button_frame, text="❌ Çıkış", 
                            command=self.root.quit,
                            bg='#34495e', fg='white', font=('Arial', 11, 'bold'),
                            relief=tk.FLAT, padx=25, pady=12)
        exit_btn.pack(side=tk.RIGHT)
        
    def create_progress_section(self, parent):
        # İlerleme çerçevesi
        progress_frame = tk.LabelFrame(parent, text="📊 İlerleme Durumu", font=('Arial', 12, 'bold'),
                                      bg='#f0f0f0', fg='#2c3e50', padx=15, pady=15)
        progress_frame.pack(fill=tk.X, pady=(0, 15))
        
        # İlerleme çubuğu
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                           maximum=100, length=600)
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        
        # Durum etiketi
        self.status_label = tk.Label(progress_frame, textvariable=self.status_var, 
                                    font=('Arial', 11), bg='#f0f0f0', fg='#7f8c8d')
        self.status_label.pack(anchor=tk.W)
        
    def create_results_section(self, parent):
        # Sonuç çerçevesi
        results_frame = tk.LabelFrame(parent, text="📈 Sonuçlar ve İstatistikler", font=('Arial', 12, 'bold'),
                                     bg='#f0f0f0', fg='#2c3e50', padx=15, pady=15)
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Sonuç metni frame
        text_frame = tk.Frame(results_frame, bg='#f0f0f0')
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Sonuç metni
        self.results_text = tk.Text(text_frame, height=12, wrap=tk.WORD, 
                                   font=('Consolas', 10), bg='#ffffff', fg='#2c3e50',
                                   relief=tk.FLAT, padx=15, pady=15)
        
        # Scroll bar
        scrollbar = tk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=scrollbar.set)
        
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Başlangıç mesajı
        self.update_results("🎯 Excel Dosya Birleştirme ve Veri Doldurma Uygulaması!\n\n"
                           "📝 Nasıl Çalışır:\n"
                           "1. İki Excel dosyasını seçin\n"
                           "2. Ortak sütunu seçin (Computer Name gibi)\n"
                           "3. Uygulama 1. dosyadaki boş alanları 2. dosyadaki değerlerle doldurur\n"
                           "4. Sonuç: Birleştirilmiş ve tamamlanmış veri\n\n"
                           "🎨 Renk Kodları:\n"
                           "🟡 Sarı: Sadece ilk dosyada bulunan kayıtlar\n"
                           "🔴 Kırmızı: Sadece ikinci dosyada bulunan kayıtlar\n"
                           "🟢 Yeşil: Birleştirilen ve doldurulan kayıtlar\n"
                           "⚪ Renksiz: Değişiklik yapılmayan kayıtlar\n\n"
                           "💡 Örnek: Computer Name ortak, Bitlocker alanı 1. dosyada boş → 2. dosyadan doldurulur")
        
    def select_file(self, file_number):
        filetypes = [
            ('Excel dosyaları', '*.xlsx *.xls'),
            ('Tüm dosyalar', '*.*')
        ]
        
        filename = filedialog.askopenfilename(
            title=f"{file_number}. Excel dosyasını seçin",
            filetypes=filetypes
        )
        
        if filename:
            if file_number == 1:
                self.file1_path.set(filename)
                self.load_file_columns(filename, 1)
            else:
                self.file2_path.set(filename)
                self.load_file_columns(filename, 2)
                
    def select_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Çıktı dosyasını kaydet",
            defaultextension=".xlsx",
            filetypes=[('Excel dosyaları', '*.xlsx'), ('Tüm dosyalar', '*.*')]
        )
        
        if filename:
            self.output_path.set(filename)
            
    def clear_all(self):
        self.file1_path.set("")
        self.file2_path.set("")
        self.output_path.set("birlestirilmis_dosya.xlsx")
        self.file1_columns = []
        self.file2_columns = []
        self.file1_column_var.set("")
        self.file2_column_var.set("")
        self.update_column_dropdowns()
        self.progress_var.set(0)
        self.status_var.set("Hazır")
        self.update_results("🗑️ Tüm alanlar temizlendi.\n\nYeni dosyalar seçebilirsiniz.")
        
    def update_results(self, text):
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, text)
        self.results_text.config(state=tk.DISABLED)
        self.results_text.see(tk.END)
        
    def update_progress(self, value, status):
        self.progress_var.set(value)
        self.status_var.set(status)
        self.root.update_idletasks()
        
    def load_file_columns(self, file_path, file_number):
        """Excel dosyasından sütun isimlerini yükle"""
        try:
            # Sadece ilk satırı oku (başlıklar)
            df = pd.read_excel(file_path, nrows=0)
            columns = list(df.columns)
            
            if file_number == 1:
                self.file1_columns = columns
            else:
                self.file2_columns = columns
                
            self.update_column_dropdowns()
            
            # Otomatik eşleştirme dene
            self.auto_match_columns()
            
        except Exception as e:
            messagebox.showwarning("⚠️ Uyarı", f"Dosya {file_number} sütunları yüklenemedi:\n{str(e)}")
            
    def update_column_dropdowns(self):
        """Dropdown menüleri güncelle"""
        if hasattr(self, 'file1_column_combo'):
            self.file1_column_combo['values'] = self.file1_columns
        if hasattr(self, 'file2_column_combo'):
            self.file2_column_combo['values'] = self.file2_columns
        
    def auto_match_columns(self):
        """Benzer sütun isimlerini otomatik eşleştir"""
        if not self.file1_columns or not self.file2_columns:
            return
            
        # Öncelikli arama kelimeleri
        priority_keywords = ['computer', 'name', 'hostname', 'pc', 'device', 'machine', 'id', 'identifier']
        
        best_match1 = None
        best_match2 = None
        
        # Önce tam eşleşme ara
        common_columns = set([col.lower() for col in self.file1_columns]) & set([col.lower() for col in self.file2_columns])
        if common_columns:
            # En öncelikli ortak sütunu seç
            for keyword in priority_keywords:
                for col in common_columns:
                    if keyword in col.lower():
                        # Orjinal isimlerini bul
                        for c1 in self.file1_columns:
                            if c1.lower() == col:
                                best_match1 = c1
                                break
                        for c2 in self.file2_columns:
                            if c2.lower() == col:
                                best_match2 = c2
                                break
                        break
                if best_match1 and best_match2:
                    break
        
        # Eğer tam eşleşme bulunamazsa, benzer isimleri ara
        if not best_match1 or not best_match2:
            for keyword in priority_keywords:
                if not best_match1:
                    for col in self.file1_columns:
                        if keyword in col.lower():
                            best_match1 = col
                            break
                
                if not best_match2:
                    for col in self.file2_columns:
                        if keyword in col.lower():
                            best_match2 = col
                            break
                
                if best_match1 and best_match2:
                    break
        
        # Bulunan eşleşmeleri ayarla
        if best_match1:
            self.file1_column_var.set(best_match1)
        elif self.file1_columns:
            self.file1_column_var.set(self.file1_columns[0])  # İlk sütunu varsayılan yap
            
        if best_match2:
            self.file2_column_var.set(best_match2)
        elif self.file2_columns:
            self.file2_column_var.set(self.file2_columns[0])  # İlk sütunu varsayılan yap
        
    def merge_files(self):
        # Dosya kontrolü
        if not self.file1_path.get() or not self.file2_path.get():
            messagebox.showerror("❌ Hata", "Lütfen her iki Excel dosyasını da seçin!")
            return
            
        if not os.path.exists(self.file1_path.get()):
            messagebox.showerror("❌ Hata", "1. dosya bulunamadı!")
            return
            
        if not os.path.exists(self.file2_path.get()):
            messagebox.showerror("❌ Hata", "2. dosya bulunamadı!")
            return
            
        # Sütun seçimi kontrolü
        if not self.file1_column_var.get() or not self.file2_column_var.get():
            messagebox.showerror("❌ Hata", "Lütfen her iki dosya için de birleştirme sütununu seçin!")
            return
            
        # Çıktı dosyası kontrolü
        if not self.output_path.get():
            messagebox.showerror("❌ Hata", "Lütfen çıktı dosyası belirtin!")
            return
            
        # Butonu devre dışı bırak
        self.merge_button.config(state=tk.DISABLED, text="⏳ İşleniyor...")
        
        # İşlemi ayrı thread'de çalıştır
        thread = threading.Thread(target=self.merge_files_thread)
        thread.daemon = True
        thread.start()
        
    def merge_files_thread(self):
        try:
            self.update_progress(10, "📁 Dosyalar okunuyor...")
            
            # Excel dosyalarını oku
            df1 = pd.read_excel(self.file1_path.get())
            df2 = pd.read_excel(self.file2_path.get())
            
            self.update_progress(25, "🔍 Dosyalar kontrol ediliyor...")
            
            # Seçilen sütunları al
            column1 = self.file1_column_var.get()
            column2 = self.file2_column_var.get()
            
            # Sütunları kontrol et
            if column1 not in df1.columns:
                raise ValueError(f"1. dosyada '{column1}' sütunu bulunamadı!")
            
            if column2 not in df2.columns:
                raise ValueError(f"2. dosyada '{column2}' sütunu bulunamadı!")
            
            self.update_progress(40, "🔧 Veriler hazırlanıyor...")
            
            # Seçilen sütunları string'e çevir
            df1[column1] = df1[column1].astype(str)
            df2[column2] = df2[column2].astype(str)
            
            # Eğer sütun isimleri farklıysa, ikincisini birinciye uyarla
            if column1 != column2:
                df2 = df2.rename(columns={column2: column1})
                column2 = column1  # Birleştirme için aynı isim kullan
            
            self.update_progress(60, "🔗 Veriler birleştiriliyor ve dolduruluyor...")
            
            # Ana dosya olarak df1'i kullan, df2'den eksik verileri doldur
            merged_df = df1.copy()
            
            # df2'den veri doldurma sayacı
            filled_count = 0
            new_records_count = 0
            
            # Her df2 kaydı için
            for _, row2 in df2.iterrows():
                key_value = row2[column1]
                
                # df1'de bu key var mı?
                mask = merged_df[column1] == key_value
                if mask.any():
                    # Var - eksik alanları doldur
                    row_index = merged_df[mask].index[0]
                    
                    # Her sütun için kontrol et
                    for col in row2.index:
                        if col != column1:  # Anahtar sütunu atlayın
                            # df1'de bu sütun var mı?
                            if col in merged_df.columns:
                                # df1'deki değer boş mu?
                                current_value = merged_df.loc[row_index, col]
                                if pd.isna(current_value) or str(current_value).strip() == '' or str(current_value).lower() == 'nan':
                                    # df2'deki değer dolu mu?
                                    new_value = row2[col]
                                    if pd.notna(new_value) and str(new_value).strip() != '' and str(new_value).lower() != 'nan':
                                        merged_df.loc[row_index, col] = new_value
                                        filled_count += 1
                            else:
                                # Bu sütun df1'de yok, ekle
                                if col not in merged_df.columns:
                                    merged_df[col] = None
                                merged_df.loc[row_index, col] = row2[col]
                                filled_count += 1
                else:
                    # Yok - yeni kayıt ekle
                    merged_df = pd.concat([merged_df, row2.to_frame().T], ignore_index=True)
                    new_records_count += 1
            
            # İstatistikleri hesapla
            total_records = len(merged_df)
            original_records = len(df1)
            
            # Excel dosyasına kaydet
            merged_df.to_excel(self.output_path.get(), index=False)
            
            self.update_progress(90, "🎨 Renkler uygulanıyor...")
            
            # Renklendirme
            wb = openpyxl.load_workbook(self.output_path.get())
            ws = wb.active
            
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Birleştirme sütununun indeksini bul
            merge_column_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == column1:
                    merge_column_col = col
                    break
            
            if merge_column_col:
                for row in range(2, ws.max_row + 1):
                    row_data = merged_df.iloc[row-2]
                    
                    file1_columns = [col for col in merged_df.columns if col.endswith('_File1')]
                    file2_columns = [col for col in merged_df.columns if col.endswith('_File2')]
                    
                    has_file1_data = any(pd.notna(row_data[col]) for col in file1_columns)
                    has_file2_data = any(pd.notna(row_data[col]) for col in file2_columns)
                    
                    if not has_file1_data and has_file2_data:
                        ws.cell(row=row, column=merge_column_col).fill = red_fill
                    elif has_file1_data and not has_file2_data:
                        ws.cell(row=row, column=merge_column_col).fill = yellow_fill
            
            wb.save(self.output_path.get())
            
            self.update_progress(100, "✅ Tamamlandı!")
            
            # Sonuçları güncelle
            results_text = f"""✅ Veri Birleştirme ve Doldurma İşlemi Tamamlandı!

📊 DETAYLI İSTATİSTİKLER:
{'='*50}
📁 Ana dosya (1. dosya) kayıt sayısı: {original_records:,}
📁 Kaynak dosya (2. dosya) kayıt sayısı: {len(df2):,}
📊 Toplam sonuç kayıt sayısı: {total_records:,}

🔗 BİRLEŞTİRME SÜTUNLARI:
{'='*50}
📄 Dosya 1 sütunu: {self.file1_column_var.get()}
📄 Dosya 2 sütunu: {self.file2_column_var.get()}

📋 İŞLEM DETAYLARI:
{'='*50}
🟢 Doldurulan alan sayısı: {filled_count:,}
🟡 Eklenen yeni kayıt sayısı: {new_records_count:,}
⚪ Değişmeden kalan kayıt sayısı: {original_records - filled_count:,}

💾 ÇIKTI DOSYASI:
{'='*50}
📄 Dosya: {os.path.basename(self.output_path.get())}
📂 Konum: {os.path.dirname(self.output_path.get())}

🎨 RENK KODLARI:
{'='*50}
🟡 Sarı: 2. dosyadan eklenen yeni kayıtlar
🟢 Yeşil: Veri doldurma işlemi yapılan kayıtlar
⚪ Renksiz: Değişiklik yapılmayan kayıtlar

💡 ÇALIŞMA PRENSİBİ:
{'='*50}
• Ana dosya (1. dosya) temel alınır
• Boş alanlar 2. dosyadan doldurulur
• 2. dosyada olup 1. dosyada olmayan kayıtlar eklenir
• Mevcut veriler korunur, sadece boş alanlar doldurulur

⚡ İşlem başarıyla tamamlandı! Excel dosyanızı açabilirsiniz.
"""
            
            self.update_results(results_text)
            
            # Başarı mesajı
            messagebox.showinfo("🎉 Başarılı", 
                              f"Veri birleştirme ve doldurma işlemi tamamlandı!\n\n"
                              f"📄 Çıktı dosyası: {os.path.basename(self.output_path.get())}\n\n"
                              f"🔗 Birleştirme sütunları:\n"
                              f"• Ana dosya: {self.file1_column_var.get()}\n"
                              f"• Kaynak dosya: {self.file2_column_var.get()}\n\n"
                              f"📊 İşlem Sonuçları:\n"
                              f"• Toplam kayıt: {total_records:,}\n"
                              f"• Doldurulan alan: {filled_count:,}\n"
                              f"• Eklenen yeni kayıt: {new_records_count:,}\n\n"
                              f"💡 1. dosyadaki boş alanlar 2. dosyadan dolduruldu!")
            
        except Exception as e:
            self.update_progress(0, "❌ Hata oluştu!")
            error_msg = f"❌ HATA OLUŞTU:\n{'='*30}\n{str(e)}\n\n💡 Lütfen dosyalarınızı kontrol edin ve tekrar deneyin."
            self.update_results(error_msg)
            messagebox.showerror("❌ Hata", str(e))
            
        finally:
            # Butonu tekrar etkinleştir
            self.merge_button.config(state=tk.NORMAL, text="🔄 Dosyaları Birleştir")
            
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    try:
        app = SimpleExcelMerger()
        app.run()
    except Exception as e:
        print(f"Uygulama başlatılamadı: {e}")
        print("Lütfen gerekli kütüphaneleri yükleyin:")
        print("pip install pandas openpyxl")
        input("Çıkmak için Enter tuşuna basın...")
